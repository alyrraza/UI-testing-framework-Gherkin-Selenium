# excel_loader.py
# Excel file se test data padhna
# Real world: Excel sheet se employee list padhna

import os
import openpyxl

# Excel file ka path
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "test_data.xlsx")

def get_login_data():
    # Excel se login test data lao
    # workbook = poori Excel file
    # worksheet = ek sheet
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook["Login"]

    data = []
    # pehli row headers hain — skip karo
    # row 2 se shuru karo
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # agar row khali nahi
            data.append({
                "username": row[0],
                "password": row[1],
                "expected": row[2]
            })
    return data

def get_checkout_data():
    # Excel se checkout test data lao
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook["Checkout"]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data.append({
                "first_name": row[0],
                "last_name": row[1],
                "zip_code": str(row[2])
            })
    return data